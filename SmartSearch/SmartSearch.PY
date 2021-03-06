﻿import tkinter as tk
import os
import datetime

'''SmartSearch is an interactive python-scripted program, that’ll help to identify the files
that matches the keyword/string as entered by user. The program looks up for the string in all
the files within the desired directory (including all the sub-folders), and will generate a
file 'SearchResult.txt' in folder 'SmartSearch' at the root directory having the list of all the
files (with their complete path) that will match the string, and also the list of files that the
program will not be able to read, due to reasons like unsupported format, or access denied, within
the same file post the matching files’ list.  '''


#-------------------------------------------------------------------------------------------------------------------------------------
# Ensures that the folder SmartSearch is present at the root Directory. Example, 'C:' in Windows OS.
#-------------------------------------------------------------------------------------------------------------------------------------
try:
    os.mkdir("C:\SmartSearch")
except:
    pass
#-------------------------------------------------------------------------------------------------------------------------------------



#-------------------------------------------------------------------------------------------------------------------------------------
# Checks if the Modules for DOCX, XLSX and XLS  files are installed, and imports them if present.
#-------------------------------------------------------------------------------------------------------------------------------------
docxFlag=0
openPyxlFlag=0
xlrdFlag=0

try:
    import docx
except ModuleNotFoundError:
    docxFlag=1

try:
    import openpyxl
except ModuleNotFoundError:
    openPyxlFlag=1

try:
    import xlrd
except ModuleNotFoundError:
    xlrdFlag=1
#-------------------------------------------------------------------------------------------------------------------------------------



#-------------------------------------------------------------------------------------------------------------------------------------
# Functions using additionally installed modules to read XLSX, XLS and DOCX files.
#-------------------------------------------------------------------------------------------------------------------------------------
def readXls(filename,str1):
    wb = xlrd.open_workbook(filename)
    sheetnames = wb.sheet_names()
    for i in sheetnames:
        sheet = wb.sheet_by_name(i)
        for row in range(sheet.nrows):
            for column in range(sheet.ncols):
                if str1 in str(sheet.cell(row,column).value).lower():
                    return 0
    return 1


def readXlsx(filename,str1):
    wb=openpyxl.load_workbook(filename)
    sheetnames=wb.get_sheet_names()
    for i in range(0,len(sheetnames)):
        sheet=wb.get_sheet_by_name(sheetnames[i])
        for j in range(1,sheet.max_row+1):
            for k in range(1,sheet.max_column+1):
                if str1 in str(sheet.cell(row=j,column=k).value).lower():
                    return 0
    return 1


def readDocx(filename,str1):
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
    a='\n'.join(fullText)
    if str1 in a.lower():
        return 0
    else:
        return 1
#-------------------------------------------------------------------------------------------------------------------------------------



def stringSearch(): #Calls searchString()
    m=searchString()


#-------------------------------------------------------------------------------------------------------------------------------------
# Function to search for user entered string
#-------------------------------------------------------------------------------------------------------------------------------------
def searchString():
    str1=e2.get().lower()
    cwd=e1.get()
    if cwd=='': #to ensure Folder Location is not Blank
        prompt = tk.Tk()
        prompt.title("Error")
        msg = tk.Label(prompt, text = "No Folder Location Found! Please enter a Folder Location value to proceed.")
        msg.config(font=(None, 12))
        msg.pack(fill='x')
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()
        return 0
    try:
        os.chdir(cwd)
    except: #to ensure that Folder Location is not invalid
        prompt = tk.Tk()
        prompt.title("Invalid Path!")
        msg = tk.Label(prompt, text = "Invalid Path! Please ensure that the path exists, or you have sufficient permission to access it.")
        msg.config(font=(None, 12))
        msg.pack(fill='x')
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()
        return 0
    if str1=='': #to ensure that search string value is not blank
        prompt = tk.Tk()
        prompt.title("Error")
        msg = tk.Label(prompt, text = "No Search String Found! Please enter a Search String value to proceed.")
        msg.config(font=(None, 12))
        msg.pack(fill='x')
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()
        return 0
    totFileCount=0
    f= open("C:\\SmartSearch\\SearchResult.txt","w+")
    count1=0 #keeps count of files that have contains the Search String
    fail=0 #keeps count of files that could not be read
    failList=[] #keeps filename and location of the files that could not be read
    f.write("You Searched For the String '"+e2.get()+"' at the path '"+cwd+"' at time "+str(datetime.datetime.now())+" !\n\n\nFiles matching the criteria:\n----------------------------\n")
    for root, dirs, files in os.walk(cwd, topdown=True):
        for name in files:
            totFileCount+=1
            x=os.path.join(root,name)
            if not os.path.isfile(x):
                continue
            try:
                if (".docx" in x.lower()) and docxFlag==0:
                    if readDocx(x,str1)==0:
                        count1+=1
                        f.write(str(count1)+") "+os.path.abspath(x)+'\n')
                    continue
                if (".xlsx" in x.lower()) and openPyxlFlag==0:
                    if readXlsx(x,str1)==0:
                        count1+=1
                        f.write(str(count1)+") "+os.path.abspath(x)+'\n')
                    continue
                if (".xls" in x.lower()) and xlrdFlag==0:
                    if readXls(x,str1)==0:
                        count1+=1
                        f.write(str(count1)+") "+os.path.abspath(x)+'\n')
                    continue
                file=open(x,"r")
                file1=file.read()
            except:
                failList.append(os.path.abspath(x))
                fail+=1
                continue
            finally:
                file.close()
            if str1 in file1.lower():
                count1+=1
                f.write(str(count1)+") "+os.path.abspath(x)+'\n')
    if count1==0:
        f.write("No Match Found!")
    if fail>0:
        f.write("\n\n\nAccess Denied or Unsupported Files : \n-----------------------------------\n")
        for ff in failList:
            f.write(ff+"\n")
    f.write("\n\n\nSearch Summary:\n---------------------\n\n"+str(count1)+" files matched the criteria!\n\nTotal Files In The Folder (Including Subfolders): "+str(totFileCount)+"\n\nFiles Unable To Read: "+str(fail))
    f.close()
    
    if count1==0: #to prompt if no file contains the Search String value
        prompt = tk.Tk()
        prompt.title("Result")
        msg = tk.Label(prompt, text = "No Match Found!")
        msg.config(font=(None, 12))
        msg.pack(ipady=4,ipadx=40)
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()
    else: #to prompt when the search is complete, and atleast 1 file contains the Search String value
        prompt = tk.Tk()
        prompt.title("Done")
        msg = tk.Label(prompt, text = "Search Complete! "+str(count1)+" out of a total "+str(totFileCount)+" files matched the search criteria.")
        msg.config(font=(None, 12))
        msg.pack(ipady=4,ipadx=40)
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()
#-------------------------------------------------------------------------------------------------------------------------------------


def readResult(): #Reads the Last Search Result
    if 'SearchResult.txt' in os.listdir('C:\\SmartSearch'):
        os.system('C:\\SmartSearch\\SearchResult.txt')
    else:
        prompt = tk.Tk()
        prompt.title("File Not Found")
        msg = tk.Label(prompt, text = "No previous search result found! Please proceed with a search to get search result.")
        msg.config(font=(None, 12))
        msg.pack(ipady=4,padx=40)
        tk.Button(prompt,text="OK",width=10,command=prompt.destroy).pack(pady=4,padx=4)
        prompt.mainloop()



#-------------------------------------------------------------------------------------------------------------------------------------
# SmartSearch main-form layout
#-------------------------------------------------------------------------------------------------------------------------------------

box1=tk.Tk()
box1.title("Welcome To SmartSearch")
loc=tk.Label(box1, text="Folder Location")
string=tk.Label(box1, text="Search String")

e1 = tk.Entry(box1, width=50)
e2 = tk.Entry(box1, width=50)
e1.insert(10,os.getcwd())


loc.grid(row=1)
e1.grid(row=1,column=1)

string.grid(row=2)
e2.grid(row=2,column=1)


x=tk.Button(box1, text='Search', width=10,command=stringSearch)
x.grid(row=3,pady=4)
x=tk.Button(box1, text='Exit',width=10, command=box1.destroy)
x.grid(row=3,column=1,sticky=tk.W,pady=4)
x=tk.Button(box1, text="Show Result",width=10,command=readResult)
x.grid(row=3,column=2,pady=4,padx=4)

box1.mainloop()
#-------------------------------------------------End Of Code---------------------------------------------------------------
